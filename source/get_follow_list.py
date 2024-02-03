from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from bs4 import BeautifulSoup
import re
import os
import time
from openpyxl import load_workbook
from source.common import page_down, find_unread_element
from source import verfication
from source import naver_script
import random


# 따로 만든 파일에서 인플루언서 아이디 목록 가져오기
# 지원 파일: txt, excel
def read_file(file_path):
    # default follow list file path: .follow_list.txt
    if file_path == "":
        file_path = "follow_list.txt"
        msg = "-i 옵션으로 파일 경로를 지정하지 않으면 기본 'follow_list.txt' 파일을 참조합니다."
        print(msg)

    # txt
    if file_path.endswith(".txt"):
        with open(file_path, "r") as f:
            influencer_list = f.readlines()
        return influencer_list

    # excel
    if file_path.endswith(".xlsx"):
        wb = load_workbook(file_path)
        sheet_list = wb.sheetnames
        ws = wb[sheet_list[0]]

        excel_col = os.environ.get("INMN_EXCEL_COL")
        if excel_col == "":
            excel_col = "B"
            msg = "엑셀의 특정 열을 지정하지 않으면 기본 엑셀 파일의 B열을 참조합니다."
            print(msg)

        influencer_list = [cell.value for cell in ws[excel_col]]
        return influencer_list


# 네이버 톡톡 목록 가져오기
# 네이버 톡톡 리스트에서 각 "https://in.naver.com" 링크 파싱
def talktalk(driver, my_influencer_id):
    influencer_list = list()

    talktalk_page = f"https://in.naver.com/{my_influencer_id}/talktalkList"
    driver.get(talktalk_page)
    time.sleep(1)

    # 접근 권한 확인: 전체/안읽은 탭 메뉴가 있는지 확인
    stat, unread_element = find_unread_element(driver)
    if stat == -1:
        print(f"{my_influencer_id}: 본인의 인플루언서 아이디를 입력해주세요. 접근 권한이 없습니다.")
        return influencer_list
    # 톡톡 목록 중 '안 읽음' 목록만 가져와서 '팬하기'를 사용할 것인지 선택
    # unread_element.click()

    # max_page만큼 PageDOWN 실행
    max_page = os.environ.get("INMN_MAX_PAGE")
    page_down(driver, max_page)

    # 톡톡 리스트에서 "https://in.naver.com" 링크 파싱
    page_html = driver.page_source
    soup = BeautifulSoup(page_html, "lxml")
    link_class_name = "TalkTalkList__ell___anpyL"
    talktalk_list = soup.find_all("span", attrs={"class": link_class_name})

    talktalk_data = ""
    for idx in range(0, len(talktalk_list), 2):
        # influencer_name = talktalk_list[idx].text
        talktalk = talktalk_list[idx+1].text
        talktalk_data += f"{talktalk}\n"

    pat = r"(?:https:\/\/in\.naver\.com\/)(\S+)"
    influencer_id_pat = re.compile(pat, re.MULTILINE)
    influencer_list = influencer_id_pat.findall(talktalk_data)

    return influencer_list


# 팬하기 네트워크:
# 나를 팬 한 사람들 중 팬 많은 순으로 정렬해서,
# 해당 인플루언서의 최신순 기준으로 "나를 팬 한 한" 목록을 뽑음
# <to-do> 상호 팬하기가 되어있는 인플루언서 목록을 가지고 연관 분석 그래프 그리기
def follower(driver, my_influencer_id):
    influencer_list = list()

    my_fan_page = f"https://in.naver.com/{my_influencer_id}/myFan"
    driver.get(my_fan_page)
    time.sleep(1)

    # 최신순/팬 많은 순 버튼이 정상적으로 로드되었는지 확인 후 클릭
    sortby_fans_xpath = "/html/body/div/div[1]/div/div[2]/div[3]/button[2]"
    stat, sortby_fans_btn = verfication.find_element(driver, By.XPATH, sortby_fans_xpath)
    if stat == -1:
        print(f"{my_influencer_id}: 본인의 인플루언서 아이디를 입력해주세요. 접근 권한이 없습니다.")
        return influencer_list
    sortby_fans_btn.click()

    # 내 팬하기 목록에서 팬 많은 순으로 팬 목록 정렬
    myfan_info_class = "MySubscriptionItem__info___xbQBT"
    myfan_info_elements = driver.find_elements(By.CLASS_NAME, myfan_info_class)

    # 일단은 랜덤으로 아무나 골라 잡아서 인플루언서 페이지로 이동
    rand_num = random.randint(0, len(myfan_info_elements)-1)
    myfan_info_elements[rand_num].send_keys(Keys.ENTER)

    # 해당 인플루언서의 팬 페이지로 이동
    follower_influencer_id = driver.current_url.split("/")[-1]
    print(follower_influencer_id)
    follower_fan_page = f"https://in.naver.com/{follower_influencer_id}/myFan"
    driver.get(follower_fan_page)
    time.sleep(1)

    # 여기서 최신순 팬 목록 추출
    max_page = os.environ.get("INMN_MAX_PAGE")
    page_down(driver, max_page)
    time.sleep(1)

    nickname_class = "MySubscriptionItem__nickname___x0Lhn"
    follower_nicknames = [elem.text for elem in driver.find_elements(By.CLASS_NAME, nickname_class)]
    for follower_nickname in follower_nicknames:
        # 인플루언서 아이디 추출
        span_text_find = f"//span[contains(text(), '{follower_nickname}')]"
        stat, follower_info_element = verfication.find_element(driver, By.XPATH, span_text_find)
        if stat == -1:
            driver.back()
            time.sleep(1)

        parent_element = follower_info_element.find_element(By.XPATH, "..").find_element(By.XPATH, "..")
        parent_element.send_keys(Keys.ENTER)
        time.sleep(0.5)
        influencer_id = driver.current_url.split("/")[-1]
        influencer_list.append(influencer_id)

        # 헤당 인플루언서 팬하기 수행
        naver_script.influencer_follow_one(driver, influencer_id)

        # 톡톡 메시지는 팬하기를 한 인플루언서에게만 전송 가능
        naver_script.send_talk(driver)
        driver.back()

        driver.back()
        time.sleep(1)

    return influencer_list
